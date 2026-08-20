from __future__ import annotations

import io
from pathlib import Path

import pytest
from openpyxl import Workbook

from swingvision_import.config import ImportConfig
from tests.swingvision_import.conftest import add_settings_and_shots_sheets
from webapp.app import create_app
from webapp.config import WebAppConfig


def _build_non_pro_workbook() -> Workbook:
    """Mirrors tests/swingvision_import/conftest.py's synthetic_non_pro_xlsx
    shape, but saved to an in-memory buffer instead of a tmp_path file, since
    the webapp needs upload bytes, not a filesystem path."""
    workbook = Workbook()

    sets_sheet = workbook.active
    sets_sheet.title = "Sets"
    sets_sheet.append(["Set", "Host Score", "Guest Score", "Set Winner"])
    sets_sheet.append([1, 6, 4, "host"])

    workbook.create_sheet("Games").append(["Game", "Set", "Server", "Game Winner"])
    workbook.create_sheet("Points").append(
        ["Point", "Game", "Set", "Match Server", "Point Winner", "Detail"]
    )

    # Point 1: host aces. Point 2 is a gap (no shots at all). Point 3: host
    # serves, guest nets a return - host wins.
    shot_rows = [
        [1, 1, "Test Player", "first_serve", "Serve", "In"],
        [3, 1, "Test Player", "first_serve", "Serve", "In"],
        [3, 2, "Test Opponent", "first_return", "Backhand", "Net"],
    ]
    add_settings_and_shots_sheets(workbook, shot_rows=shot_rows)
    return workbook


@pytest.fixture
def xlsx_bytes() -> bytes:
    buffer = io.BytesIO()
    _build_non_pro_workbook().save(buffer)
    return buffer.getvalue()


@pytest.fixture
def import_config(tmp_path: Path) -> ImportConfig:
    return ImportConfig(pending_dir=tmp_path / "pending", db_path=tmp_path / "court_iq.db")


@pytest.fixture
def webapp_config(tmp_path: Path) -> WebAppConfig:
    return WebAppConfig(uploads_dir=tmp_path / "uploads", media_dir=tmp_path / "media")


@pytest.fixture
def app(import_config: ImportConfig, webapp_config: WebAppConfig):
    flask_app = create_app(import_config, webapp_config)
    flask_app.config["TESTING"] = True
    return flask_app


@pytest.fixture
def client(app):
    return app.test_client()
