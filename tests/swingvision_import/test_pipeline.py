from __future__ import annotations

import json
import logging
import sqlite3

import pytest
from openpyxl import Workbook

from swingvision_import.load import UnresolvedReviewError
from swingvision_import.pipeline import SwingVisionImportPipeline
from swingvision_import.records import MatchRecord
from swingvision_import.review import load_pending, save_pending
from tests.ai.conftest import FakeMessage, FakeTextBlock


def _clear_review_flags(pipeline, json_path):
    record = load_pending(json_path)
    for set_record in record.sets:
        for point in set_record.points:
            point.needs_review = False
    save_pending(record, pipeline.config.pending_dir)


def test_ingest_writes_pending_json_without_touching_sql(synthetic_xlsx, import_config):
    pipeline = SwingVisionImportPipeline(import_config)
    json_path = pipeline.ingest(synthetic_xlsx, date="2026-08-06", opponent="Alex", result="W")

    assert json_path.exists()
    assert not import_config.db_path.exists()


def test_finalize_rejects_until_flags_resolved(synthetic_xlsx, import_config):
    pipeline = SwingVisionImportPipeline(import_config)
    json_path = pipeline.ingest(synthetic_xlsx, date="2026-08-06", opponent="Alex", result="W")

    with pytest.raises(UnresolvedReviewError):
        pipeline.finalize(json_path)


def test_finalize_loads_once_flags_are_cleared(synthetic_xlsx, import_config):
    pipeline = SwingVisionImportPipeline(import_config)
    json_path = pipeline.ingest(synthetic_xlsx, date="2026-08-06", opponent="Alex", result="W")
    _clear_review_flags(pipeline, json_path)

    match_id = pipeline.finalize(json_path)
    assert match_id == 1


def test_rerunning_finalize_on_same_match_is_rejected_not_duplicated(synthetic_xlsx, import_config):
    pipeline = SwingVisionImportPipeline(import_config)
    json_path = pipeline.ingest(synthetic_xlsx, date="2026-08-06", opponent="Alex", result="W")
    _clear_review_flags(pipeline, json_path)

    pipeline.finalize(json_path)
    with pytest.raises(ValueError):
        pipeline.finalize(json_path)


def test_match_overrides_survive_the_full_ingest_then_finalize_round_trip(
    synthetic_xlsx, import_config
):
    pipeline = SwingVisionImportPipeline(import_config)
    json_path = pipeline.ingest(
        synthetic_xlsx,
        date="2026-08-06",
        opponent="Alex",
        result="W",
        energy_rating=5,
        pros="Aggressive on return games",
    )
    _clear_review_flags(pipeline, json_path)
    pipeline.finalize(json_path)

    connection = sqlite3.connect(import_config.db_path)
    row = connection.execute("SELECT energy_rating, pros FROM match").fetchone()
    connection.close()
    assert row == (5, "Aggressive on return games")


def test_reingesting_the_same_match_overwrites_the_pending_file_at_the_same_path(
    synthetic_xlsx, import_config
):
    pipeline = SwingVisionImportPipeline(import_config)
    first_path = pipeline.ingest(synthetic_xlsx, date="2026-08-06", opponent="Alex", result="W")
    second_path = pipeline.ingest(
        synthetic_xlsx, date="2026-08-06", opponent="Alex", result="W", pros="Revised notes"
    )

    assert first_path == second_path
    assert list(import_config.pending_dir.glob("*.json")) == [first_path]


def test_ingest_falls_back_to_shot_reconstruction_when_points_sheet_is_empty(
    synthetic_non_pro_xlsx, import_config
):
    pipeline = SwingVisionImportPipeline(import_config)

    json_path = pipeline.ingest(
        synthetic_non_pro_xlsx, date="2026-08-06", opponent="Alex", result="W"
    )
    record = load_pending(json_path)

    all_points = [p for s in record.sets for p in s.points]
    assert len(all_points) == 2
    assert all(p.needs_review for p in all_points)
    assert all(p.source_point_number is not None for p in all_points)
    # Point 1 was a single In serve shot by the host - an ace.
    assert any(p.point_end_type == "ace" for p in all_points)


class _FakeSuggestionMessages:
    def __init__(self):
        self.calls = []

    def create(self, **kwargs):
        self.calls.append(kwargs)
        payload = json.dumps(
            {"point_end_type": "forced_error", "reasoning": "Deep shot.", "confidence": "medium"}
        )
        return FakeMessage(content=[FakeTextBlock(text=payload)])


class _FakeSuggestionClient:
    def __init__(self):
        self.messages = _FakeSuggestionMessages()


def test_suggest_annotates_reconstructed_points_without_touching_needs_review(
    synthetic_non_pro_xlsx, import_config
):
    pipeline = SwingVisionImportPipeline(import_config)
    json_path = pipeline.ingest(
        synthetic_non_pro_xlsx, date="2026-08-06", opponent="Alex", result="W"
    )
    client = _FakeSuggestionClient()

    record = pipeline.suggest(client, json_path)

    all_points = [p for s in record.sets for p in s.points]
    assert all(p.ai_suggested_point_end_type == "forced_error" for p in all_points)
    assert all(p.ai_suggestion_reasoning == "Deep shot." for p in all_points)
    assert all(p.needs_review for p in all_points)  # never cleared
    assert len(client.messages.calls) == len(all_points)

    # Suggestions persist to the same pending file.
    reloaded = load_pending(json_path)
    assert reloaded.sets[0].points[0].ai_suggested_point_end_type == "forced_error"


def test_suggest_raises_without_a_source_file_to_re_parse(import_config):
    pipeline = SwingVisionImportPipeline(import_config)
    record = MatchRecord(date="2026-08-06", opponent="Alex", result="W")
    json_path = save_pending(record, import_config.pending_dir)

    with pytest.raises(ValueError, match="source_file"):
        pipeline.suggest(_FakeSuggestionClient(), json_path)


def test_suggest_skips_direct_parse_points_with_no_shot_provenance(
    synthetic_xlsx, import_config
):
    # synthetic_xlsx has real Points-sheet rows, so ingest() takes the
    # direct-parse path - those points have no source_point_number and no
    # raw shots to reason over, so suggest() must leave them alone.
    pipeline = SwingVisionImportPipeline(import_config)
    json_path = pipeline.ingest(synthetic_xlsx, date="2026-08-06", opponent="Alex", result="W")
    client = _FakeSuggestionClient()

    record = pipeline.suggest(client, json_path)

    all_points = [p for s in record.sets for p in s.points]
    assert all(p.source_point_number is None for p in all_points)
    assert all(p.ai_suggested_point_end_type is None for p in all_points)
    assert len(client.messages.calls) == 0


def test_ingest_raises_when_points_is_empty_and_settings_is_missing(tmp_path, import_config):
    workbook = Workbook()
    sets_sheet = workbook.active
    sets_sheet.title = "Sets"
    sets_sheet.append(["Set", "Host Score", "Guest Score", "Set Winner"])
    sets_sheet.append([1, 6, 4, "host"])
    workbook.create_sheet("Games").append(["Game", "Set", "Server", "Game Winner"])
    workbook.create_sheet("Points").append(
        ["Point", "Game", "Set", "Match Server", "Point Winner", "Detail"]
    )
    # A Settings sheet header with no data row at all -> raw.settings is None.
    workbook.create_sheet("Settings").append(["Host Team", "Guest Team"])
    workbook.create_sheet("Shots").append(["Point", "Shot", "Player", "Type", "Stroke", "Result"])
    path = tmp_path / "no_settings.xlsx"
    workbook.save(path)

    pipeline = SwingVisionImportPipeline(import_config)
    with pytest.raises(ValueError, match="Settings sheet"):
        pipeline.ingest(path, date="2026-08-06", opponent="Alex", result="W")


class _FlakySuggestionMessages:
    """Fails for the first point, succeeds for the rest."""

    def __init__(self):
        self.calls = []

    def create(self, **kwargs):
        self.calls.append(kwargs)
        if len(self.calls) == 1:
            return FakeMessage(content=[FakeTextBlock(text="not valid json")])
        payload = json.dumps(
            {"point_end_type": "winner", "reasoning": "Clean pass.", "confidence": "high"}
        )
        return FakeMessage(content=[FakeTextBlock(text=payload)])


class _FlakySuggestionClient:
    def __init__(self):
        self.messages = _FlakySuggestionMessages()


def test_suggest_continues_past_one_points_suggestion_failure(
    synthetic_non_pro_xlsx, import_config, caplog
):
    pipeline = SwingVisionImportPipeline(import_config)
    json_path = pipeline.ingest(
        synthetic_non_pro_xlsx, date="2026-08-06", opponent="Alex", result="W"
    )
    client = _FlakySuggestionClient()

    with caplog.at_level(logging.ERROR):
        record = pipeline.suggest(client, json_path)

    all_points = [p for s in record.sets for p in s.points]
    succeeded = [p for p in all_points if p.ai_suggested_point_end_type == "winner"]
    failed = [p for p in all_points if p.ai_suggested_point_end_type is None]
    assert len(succeeded) == 1
    assert len(failed) == 1
    assert any("Suggestion failed" in r.message for r in caplog.records)
