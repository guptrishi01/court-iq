from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

from swingvision_import.config import ImportConfig


@pytest.fixture
def synthetic_xlsx(tmp_path: Path) -> Path:
    """A small hand-built .xlsx standing in for a real SwingVision export.

    SwingVision publishes no official export schema, so this encodes our best
    guess (see config.DEFAULT_COLUMN_ALIASES). Replace/extend once a real
    exported match is available to check the guess against.
    """
    workbook = Workbook()

    sets_sheet = workbook.active
    sets_sheet.title = "Sets"
    sets_sheet.append(["Set #", "Winner", "Games Won", "Games Lost"])
    sets_sheet.append([1, "player", 6, 4])

    games_sheet = workbook.create_sheet("Games")
    games_sheet.append(["Set #", "Game #", "Server", "Winner"])
    games_sheet.append([1, 1, "player", "player"])
    games_sheet.append([1, 2, "opponent", "opponent"])

    points_sheet = workbook.create_sheet("Points")
    points_sheet.append(
        ["Set #", "Game #", "Point #", "Server", "Winner",
         "1st Serve In", "2nd Serve In", "Shot Type"]
    )
    points_sheet.append([1, 1, 1, "player", "player", True, None, "ace"])
    points_sheet.append([1, 1, 2, "player", "player", False, True, "winner"])
    points_sheet.append([1, 2, 1, "opponent", "player", True, None, "unforced_error"])

    path = tmp_path / "synthetic_export.xlsx"
    workbook.save(path)
    return path


@pytest.fixture
def import_config(tmp_path: Path) -> ImportConfig:
    return ImportConfig(
        pending_dir=tmp_path / "pending",
        db_path=tmp_path / "court_iq.db",
    )
