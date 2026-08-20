from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

from swingvision_import.config import ImportConfig
from swingvision_import.parse import SwingVisionParser
from tests.swingvision_import.conftest import add_settings_and_shots_sheets


def test_parse_reads_sets_games_and_points(synthetic_xlsx):
    raw = SwingVisionParser(ImportConfig()).parse(synthetic_xlsx)

    assert len(raw.sets) == 1
    assert raw.sets[0].games_won == 6
    assert raw.sets[0].games_lost == 4

    assert len(raw.games) == 2
    assert raw.games[0].server == "host"
    assert raw.games[1].winner == "guest"

    assert len(raw.points) == 3
    assert raw.points[0].end_type == "ace"
    assert raw.points[0].first_serve_in is True
    assert raw.points[1].second_serve_in is True


def test_parse_reads_settings_host_and_guest_names(synthetic_xlsx):
    raw = SwingVisionParser(ImportConfig()).parse(synthetic_xlsx)

    assert raw.settings is not None
    assert raw.settings.host_name == "Test Player"
    assert raw.settings.guest_name == "Test Opponent"


def test_parse_reads_shots_and_skips_trailing_blank_rows(tmp_path: Path):
    workbook = Workbook()
    sets_sheet = workbook.active
    sets_sheet.title = "Sets"
    sets_sheet.append(["Set", "Host Score", "Guest Score", "Set Winner"])
    sets_sheet.append([1, 6, 4, "host"])
    workbook.create_sheet("Games").append(["Game", "Set", "Server", "Game Winner"])
    workbook.create_sheet("Points").append(
        ["Point", "Game", "Set", "Match Server", "Point Winner", "Detail"]
    )

    shot_rows = [
        [1, 1, "Rishi Gupta", "first_serve", "Serve", "In"],
        [None, None, None, None, None, None],
    ]
    add_settings_and_shots_sheets(workbook, shot_rows=shot_rows)

    path = tmp_path / "shots_with_blank_row.xlsx"
    workbook.save(path)

    raw = SwingVisionParser(ImportConfig()).parse(path)

    assert len(raw.shots) == 1
    assert raw.shots[0].player == "Rishi Gupta"
    assert raw.shots[0].shot_type == "first_serve"


def test_parse_accepts_text_booleans_and_skips_trailing_blank_rows(tmp_path: Path):
    """Real exports may store yes/no as text rather than Excel bool cells, and
    often have a trailing blank row after the last real point."""
    workbook = Workbook()

    sets_sheet = workbook.active
    sets_sheet.title = "Sets"
    sets_sheet.append(["Set #", "Winner", "Games Won", "Games Lost"])
    sets_sheet.append([1, "host", 6, 4])

    games_sheet = workbook.create_sheet("Games")
    games_sheet.append(["Set #", "Game #", "Server", "Winner"])
    games_sheet.append([1, 1, "host", "host"])

    points_sheet = workbook.create_sheet("Points")
    points_sheet.append(
        ["Set #", "Game #", "Point #", "Server", "Winner",
         "1st Serve In", "2nd Serve In", "Shot Type"]
    )
    points_sheet.append([1, 1, 1, "host", "host", "Yes", "No", "ace"])
    points_sheet.append([None, None, None, None, None, None, None, None])

    add_settings_and_shots_sheets(workbook)

    path = tmp_path / "text_booleans.xlsx"
    workbook.save(path)

    raw = SwingVisionParser(ImportConfig()).parse(path)

    assert len(raw.points) == 1
    assert raw.points[0].first_serve_in is True
    assert raw.points[0].second_serve_in is False


def test_parse_matches_headers_case_and_whitespace_insensitively(tmp_path: Path):
    """We don't know SwingVision's real header casing/spacing, so alias
    matching should not depend on getting it exactly right."""
    workbook = Workbook()

    sets_sheet = workbook.active
    sets_sheet.title = "Sets"
    # Deliberately different case/whitespace from the documented aliases.
    sets_sheet.append([" set #  ", "WINNER", "games won", "Games Lost"])
    sets_sheet.append([1, "host", 6, 4])

    games_sheet = workbook.create_sheet("Games")
    games_sheet.append(["Set #", "Game #", "Server", "Winner"])
    games_sheet.append([1, 1, "host", "host"])

    points_sheet = workbook.create_sheet("Points")
    points_sheet.append(
        ["Set #", "Game #", "Point #", "Server", "Winner",
         "1st Serve In", "2nd Serve In", "Shot Type"]
    )
    points_sheet.append([1, 1, 1, "host", "host", True, None, "ace"])

    add_settings_and_shots_sheets(workbook)

    path = tmp_path / "case_insensitive.xlsx"
    workbook.save(path)

    raw = SwingVisionParser(ImportConfig()).parse(path)

    assert raw.sets[0].games_won == 6
    assert raw.sets[0].games_lost == 4


def test_parse_raises_a_clear_error_when_a_required_column_is_missing(tmp_path: Path):
    """A silently-mismatched header must not fall through to garbage data
    (e.g. a missing 'Winner' column stringifying None into the literal text
    'None' as the point winner) - it should fail loudly and name the sheet."""
    workbook = Workbook()

    sets_sheet = workbook.active
    sets_sheet.title = "Sets"
    sets_sheet.append(["Set #", "Winner", "Games Won", "Games Lost"])
    sets_sheet.append([1, "host", 6, 4])

    games_sheet = workbook.create_sheet("Games")
    games_sheet.append(["Set #", "Game #", "Server", "Winner"])
    games_sheet.append([1, 1, "host", "host"])

    points_sheet = workbook.create_sheet("Points")
    # "Winner" column renamed to something not in the alias list.
    points_sheet.append(
        ["Set #", "Game #", "Point #", "Server", "Point Result", "Shot Type"]
    )
    points_sheet.append([1, 1, 1, "host", "host", "ace"])

    add_settings_and_shots_sheets(workbook)

    path = tmp_path / "missing_column.xlsx"
    workbook.save(path)

    with pytest.raises(ValueError, match="Points"):
        SwingVisionParser(ImportConfig()).parse(path)
