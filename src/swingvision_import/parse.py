"""Reads a SwingVision .xlsx export into raw dataclasses.

Column positions are resolved by header name (via ImportConfig's alias map)
rather than hard-coded indices, so a real export with slightly different
headers only requires a config change, not a rewrite here.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .config import ImportConfig
from .raw import RawGameRow, RawMatchExport, RawPointRow, RawSetRow, RawSettings, RawShotRow

# Fields that may legitimately be absent from a real export (SwingVision may
# not always report a second-serve column on an ace/first-serve-in point,
# for instance). Anything else missing from the header is treated as a
# configuration error, not silently coerced into a garbage value.
_OPTIONAL_FIELDS: dict[str, frozenset[str]] = {
    "points": frozenset({"first_serve_in", "second_serve_in"}),
}


def _resolve_columns(header_row: tuple, field_aliases: dict[str, list[str]]) -> dict[str, int]:
    """Maps each canonical field name to its column index in a header row.

    Matching is case- and whitespace-insensitive, since SwingVision's real
    header casing/spacing is unverified.

    Args:
        header_row: The sheet's header row, as returned by
            Worksheet.iter_rows(values_only=True).
        field_aliases: Maps each canonical field name to the list of header
            strings that may represent it.

    Returns:
        A dict mapping each field name found in the header to its column
        index. Field names with no matching header column are omitted.
    """
    resolved: dict[str, int] = {}
    for field_name, aliases in field_aliases.items():
        normalized_aliases = {alias.strip().lower() for alias in aliases}
        for idx, cell in enumerate(header_row):
            if cell is not None and str(cell).strip().lower() in normalized_aliases:
                resolved[field_name] = idx
                break
    return resolved


def _check_required_columns(
    sheet_key: str, sheet_name: str, columns: dict[str, int], field_aliases: dict[str, list[str]]
) -> None:
    """Raises if a non-optional field has no matching column.

    Without this check, a mismatched header would silently produce garbage
    values (e.g. a missing "Winner" column stringifying None into the
    literal text "None") instead of failing loudly.

    Args:
        sheet_key: The sheet's logical key (e.g. "points"), used to look up
            which of its fields are allowed to be absent.
        sheet_name: The sheet's display name, for the error message.
        columns: The result of _resolve_columns for this sheet.
        field_aliases: The full set of field names expected for this sheet.

    Raises:
        ValueError: If one or more required fields have no matching column.
    """
    optional = _OPTIONAL_FIELDS.get(sheet_key, frozenset())
    missing = [name for name in field_aliases if name not in columns and name not in optional]
    if missing:
        raise ValueError(
            f"Could not find a column for {missing} in the '{sheet_name}' sheet "
            f"(looked for: {[field_aliases[name] for name in missing]}). "
            "Update ImportConfig.column_aliases to match this export's actual headers."
        )


def _cell(row: tuple, columns: dict[str, int], field_name: str) -> Any:
    """Reads one field's value out of a row, or None if it wasn't found.

    Args:
        row: A single data row from the sheet.
        columns: Field name -> column index, as returned by
            _resolve_columns.
        field_name: The canonical field name to read.

    Returns:
        The cell value at that field's column, or None if the field has no
        matching column in this sheet.
    """
    idx = columns.get(field_name)
    return row[idx] if idx is not None else None


def _as_bool(value: Any) -> bool | None:
    """Normalizes a cell value into a bool, tolerating text and blanks.

    Real exports may store yes/no as text ("Yes"/"No", "In"/"Out") rather
    than native Excel boolean cells.

    Args:
        value: The raw cell value.

    Returns:
        True or False if the value can be interpreted as one, or None if
        the cell was blank.
    """
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() in {"true", "1", "yes", "in"}


class SwingVisionParser:
    """Parses a SwingVision .xlsx export into raw dataclasses.

    Column positions are resolved by header name via the alias map on the
    given ImportConfig, so a real export with slightly different headers
    only requires a config change, not a change to this class.
    """

    def __init__(self, config: ImportConfig) -> None:
        """Initializes the parser.

        Args:
            config: Sheet names and column aliases to parse against.
        """
        self._config = config

    def parse(self, xlsx_path: Path) -> RawMatchExport:
        """Parses a SwingVision export file into raw dataclasses.

        Args:
            xlsx_path: Path to the exported .xlsx file.

        Returns:
            The parsed settings, sets, games, points, and shots.

        Raises:
            ValueError: If a required column can't be found in one of the
                Settings, Sets, Games, Points, or Shots sheets.
        """
        workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
        try:
            return RawMatchExport(
                settings=self._parse_settings(workbook),
                sets=self._parse_sets(workbook),
                games=self._parse_games(workbook),
                points=self._parse_points(workbook),
                shots=self._parse_shots(workbook),
            )
        finally:
            workbook.close()

    def _sheet_rows(self, workbook: Any, sheet_key: str) -> tuple[dict[str, int], Any]:
        """Resolves a sheet's header columns and returns its data rows.

        Args:
            workbook: The open openpyxl Workbook.
            sheet_key: The sheet's logical key (e.g. "points").

        Returns:
            A tuple of (field name -> column index, an iterator over the
            sheet's remaining data rows).

        Raises:
            ValueError: If a required field has no matching column.
        """
        sheet_name = self._config.sheet_names[sheet_key]
        sheet = workbook[sheet_name]
        rows = sheet.iter_rows(values_only=True)
        header = next(rows)
        field_aliases = self._config.column_aliases[sheet_key]
        columns = _resolve_columns(header, field_aliases)
        _check_required_columns(sheet_key, sheet_name, columns, field_aliases)
        return columns, rows

    def _parse_sets(self, workbook: Any) -> list[RawSetRow]:
        """Parses the "Sets" sheet.

        Args:
            workbook: The open openpyxl Workbook.

        Returns:
            One RawSetRow per non-blank row in the sheet.
        """
        columns, rows = self._sheet_rows(workbook, "sets")
        return [
            RawSetRow(
                set_number=int(_cell(row, columns, "set_number")),
                winner=str(_cell(row, columns, "winner")),
                games_won=int(_cell(row, columns, "games_won")),
                games_lost=int(_cell(row, columns, "games_lost")),
            )
            for row in rows
            if _cell(row, columns, "set_number") is not None
        ]

    def _parse_games(self, workbook: Any) -> list[RawGameRow]:
        """Parses the "Games" sheet.

        Args:
            workbook: The open openpyxl Workbook.

        Returns:
            One RawGameRow per non-blank row in the sheet.
        """
        columns, rows = self._sheet_rows(workbook, "games")
        return [
            RawGameRow(
                set_number=int(_cell(row, columns, "set_number")),
                game_number=int(_cell(row, columns, "game_number")),
                server=str(_cell(row, columns, "server")),
                winner=str(_cell(row, columns, "winner")),
            )
            for row in rows
            if _cell(row, columns, "game_number") is not None
        ]

    def _parse_settings(self, workbook: Any) -> RawSettings | None:
        """Parses the "Settings" sheet's single data row.

        Unlike the other sheets, Settings has exactly one real data row
        (match metadata), followed by blank rows and then coordinate-system
        footnote text — so this reads only the first row after the header,
        rather than looping until some column goes blank.

        Args:
            workbook: The open openpyxl Workbook.

        Returns:
            The match's host/guest names, or None if the sheet had no data
            row at all.
        """
        columns, rows = self._sheet_rows(workbook, "settings")
        row = next(rows, None)
        if row is None:
            return None
        return RawSettings(
            host_name=str(_cell(row, columns, "host_name")),
            guest_name=str(_cell(row, columns, "guest_name")),
        )

    def _parse_shots(self, workbook: Any) -> list[RawShotRow]:
        """Parses the "Shots" sheet.

        Args:
            workbook: The open openpyxl Workbook.

        Returns:
            One RawShotRow per non-blank row in the sheet.
        """
        columns, rows = self._sheet_rows(workbook, "shots")
        shots = []
        for row in rows:
            if _cell(row, columns, "point_number") is None:
                continue
            shots.append(
                RawShotRow(
                    point_number=int(_cell(row, columns, "point_number")),
                    shot_number=int(_cell(row, columns, "shot_number")),
                    player=str(_cell(row, columns, "player")),
                    shot_type=str(_cell(row, columns, "shot_type")),
                    stroke=str(_cell(row, columns, "stroke")),
                    result=str(_cell(row, columns, "result")),
                )
            )
        return shots

    def _parse_points(self, workbook: Any) -> list[RawPointRow]:
        """Parses the "Points" sheet.

        Args:
            workbook: The open openpyxl Workbook.

        Returns:
            One RawPointRow per non-blank row in the sheet. Trailing blank
            rows (common at the end of a real export) are skipped.
        """
        columns, rows = self._sheet_rows(workbook, "points")
        points = []
        for row in rows:
            if _cell(row, columns, "point_number") is None:
                continue
            points.append(
                RawPointRow(
                    set_number=int(_cell(row, columns, "set_number")),
                    game_number=int(_cell(row, columns, "game_number")),
                    point_number=int(_cell(row, columns, "point_number")),
                    server=str(_cell(row, columns, "server")),
                    winner=str(_cell(row, columns, "winner")),
                    end_type=str(_cell(row, columns, "end_type") or "").strip().lower(),
                    first_serve_in=_as_bool(_cell(row, columns, "first_serve_in")),
                    second_serve_in=_as_bool(_cell(row, columns, "second_serve_in")),
                )
            )
        return points
